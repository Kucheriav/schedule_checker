from design import Ui_MainWindow
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QLabel
from openpyxl import load_workbook, Workbook
from sche_che import FilePreparator, DifferenceEngine
import sys


class Window(QMainWindow, Ui_MainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        self.setupUi(self)
        self.comboBox.setEnabled(False)
        self.comboBox.addItems(['понедельник', 'вторник', 'среда', 'четверг', 'пятница'])
        self.openBaseScheduleBtn.clicked.connect(self.openFile)
        self.openNewScheduleBtn.clicked.connect(self.openFile)
        self.findChangesBtn.clicked.connect(self.compare_files)
        self.checkBox.clicked.connect(lambda x: self.comboBox.setEnabled(not self.comboBox.isEnabled()))
        self.work_dir = ''
        self.base_schedule = None
        self.new_schedule = None
        self.findChangesBtn.setEnabled(False)
        self.old_file_preparation_task = FilePreparator()
        self.new_file_preparation_task = FilePreparator()
        self.difference_search_task = DifferenceEngine()
        self.baseProgressBar.setValue(0)

        self.old_file_preparation_task.preparation_progress.connect(self.baseProgressBar.setValue)
        self.newProgressBar.setValue(0)
        self.new_file_preparation_task.preparation_progress.connect(self.newProgressBar.setValue)

    def openFile(self):
        filename, _ = QFileDialog.getOpenFileName(self, 'Выбрать файл',' self.work_dir', 'Excel файлы (*.xlsx)')
        self.work_dir = '/'.join(filename.split('/')[:-1])
        if filename:
            if self.sender() == self.openBaseScheduleBtn:
                self.base_schedule = load_workbook(filename)
                self.baseScheduleLabel.setText(f"Выбрано: {'/'.join(filename.split('/')[-2:])[:-5]}")
            else:
                self.new_schedule = load_workbook(filename)
                self.newScheduleLabel.setText(f"Выбрано: {'/'.join(filename.split('/')[-2:])[:-5]}")
        if self.base_schedule and self.new_schedule:
            self.findChangesBtn.setEnabled(True)

    def compare_files(self):

        self.baseProgressBar.setMaximum(self.base_schedule.active.max_row)
        self.base_schedule = self.old_file_preparation_task.row_normalization(self.base_schedule)

        self.newProgressBar.setMaximum(self.new_schedule.active.max_row)
        self.new_schedule = self.new_file_preparation_task.row_normalization(self.new_schedule)

        self.difference_search_task.bold_difference_v2(self.base_schedule, self.new_schedule)
        self.new_schedule.save(f"{self.work_dir}/{self.newScheduleLabel.text().split('/')[1]}_checked.xlsx")


def except_hook(cls, exception, traceback):
    sys.__excepthook__(cls, exception, traceback)

# обвязка для запуска
app = QApplication(sys.argv)
ex = Window()
ex.show()
sys.excepthook = except_hook
sys.exit(app.exec_())


