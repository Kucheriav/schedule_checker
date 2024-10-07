from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from PyQt5.QtCore import QObject, pyqtSignal
import argparse
from tqdm import tqdm
from openpyxl.utils import get_column_letter

from db_models import Class, Schedule, Teacher, TeacherSchedule, Cabinet
from database import get_db


DAYS = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница']
CABINETS_WITH_EL_SCHOOL = ['101', '102', '103', '107', '201', '202', '203', '204', '205', '206', '207', '208', '209', '301', '302',
            '303', '304', '305', '306', '307', '308', '401', '402', '403', '404', '405', '406', '407', '408', '409',
            '411', '412', 'Акт.зал', 'СЗ', 'СЗ', 'СЗ', 'П']
CABINETS = ['101', '102', '107', '208', '209', '301', '302', '303', '304', '305', '306', '307', '308', '401', '402',
            '403', '404', '405', '406', '407', '408', '409', '411', '412', 'Акт.зал', 'СЗ', 'СЗ', 'СЗ', 'П']

def load_data_from_excel(file_path):
    # Пример загрузки данных из Excel файла
    df = pd.read_excel(file_path)
    db = next(get_db())
    for index, row in df.iterrows():
        class_ = Class(name=row['class_name'])
        db.add(class_)
        db.commit()

        schedule = Schedule(class_id=class_.id, day=row['day'], lesson_number=row['lesson_number'], subject=row['subject'], cabinet=row['cabinet'])
        db.add(schedule)
        db.commit()
    db.close()

def export_data_to_excel(file_path):
    # Пример экспорта данных в Excel файл
    db = next(get_db())
    schedules = db.query(Schedule).all()
    df = pd.DataFrame([schedule.__dict__ for schedule in schedules])
    df.to_excel(file_path, index=False)
    db.close()



class FuncToolBox(QObject):
    progress_status = pyqtSignal(int)

    def row_normalization(self, wb):
        # ае! я поборол это дело без дублирования строк!

        # тащемта у нас всего два проблемных случая спаренной по вертикали строки
        # это когда шапка класса и собственно урок в двух кабинетах
        ws = wb.active
        wb_out = Workbook()
        ws_out = wb_out.active

        merged_cells = ws.merged_cells.ranges
        row = 1
        row_out = 1

        pbar = tqdm(total=ws.max_row)
        while row < ws.max_row:
            # делаем шапку класса
            if ws.cell(row, 1).value and ws.cell(row, 1).value == '#':
                ws_out.append([None for x in range(11)])
                ws_out.append([None for x in range(11)])
                ws_out.merge_cells(start_row=row_out, start_column=1, end_row=row_out + 1, end_column=1)
                ws_out.cell(row_out, 1).value = '№'
                for i in range(5):
                    ws_out.merge_cells(start_row=row_out, start_column=2 + i * 2, end_row=row_out,
                                       end_column=2 + i * 2 + 1)
                    ws_out.cell(row_out, 2 + i * 2).value = ws.cell(row, 2 + i * 2).value
                    ws_out.cell(row_out + 1, 2 + i * 2).value = ws.cell(row + 1, 2 + i * 2).value
                    ws_out.cell(row_out + 1, 2 + i * 2 + 1).value = ws.cell(row + 1, 2 + i * 2 + 1).value
                row += 2
                row_out += 2
                pbar.update(2)
            # случай спаренной строки урока
            elif (any(ws.cell(row, 1).coordinate in range_str for range_str in merged_cells) and ws.cell(row, 1).value
                  and not ws.cell(row, 1).value == '#' and not 'Класс' in str(ws.cell(row, 1).value)):
                this_row = ws[row]
                next_row = ws[row + 1]
                for this_row_cell, next_row_cell in zip(this_row, next_row):
                    if next_row_cell.value is not None:
                        this_row_cell.value = f'{this_row_cell.value}\n{next_row_cell.value}'
                ws_out.append([cell.value for cell in this_row])
                if ':' in str(ws_out.cell(row_out, 1).value):
                    ws_out.cell(row_out, 1).value = str(ws_out.cell(row_out, 1).value)[:5]
                row += 2
                row_out += 1
                pbar.update(2)
            else:
                #все остальные случаи не содержат объединений на вертикали, так что пофиг
                ws_out.append([cell.value for cell in ws[row]])
                if ':' in str(ws_out.cell(row_out, 1).value):
                    ws_out.cell(row_out, 1).value = str(ws_out.cell(row_out, 1).value)[1:5]
                row += 1
                row_out += 1
                pbar.update(1)
        pbar.close()
        return wb_out

    def bold_difference(self, old_wb, new_wb):
        dif_cell_font = Font(bold=True)
        old_ws = old_wb.active
        new_ws = new_wb.active
        row = 1
        old_row = 1
        while row < new_ws.max_row:
            if 'Класс' in str(new_ws.cell(row, 1).value):
                print(new_ws.cell(row, 1).value, row)
                flag = False
                for x in range(old_row, old_ws.max_row + 1):
                    if old_ws.cell(x, 1).value == new_ws.cell(row, 1).value:
                        old_row = x
                        flag = True
                        break
                if not flag:
                    print('No matches')
                    raise Exception
                cur_new_row = row + 3
                cur_old_row = old_row + 3
                while not (new_ws.cell(cur_new_row, 1).value is None):
                    for col in range(1, len(new_ws[cur_new_row]) + 1):
                        if new_ws.cell(cur_new_row, col).value != old_ws.cell(cur_old_row, col).value:
                            if new_ws.cell(cur_new_row, col).value is None:
                                print(cur_new_row, col)
                                new_ws.cell(cur_new_row, col).value = '-окно-'
                            new_ws.cell(cur_new_row, col).font = dif_cell_font
                            new_ws.cell(cur_new_row, col).fill = PatternFill(start_color='ffff00', end_color='ffff00',
                                                                             fill_type='solid')
                            print(cur_new_row, col)
                    cur_old_row += 1
                    cur_new_row += 1
                row = cur_new_row
            row += 1
        return new_wb

    def day_assemble(self, wb, day):
        res_wb = Workbook()
        res_ws = res_wb.active
        ws_in = wb.active
        res_ws.append([None])
        res_ws.append([None])
        res_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
        res_ws.cell(1, 1).value = '№'
        for i in range(1, 12):
            res_ws.append([str(i)])
        res_wb.save('test.xlsx')
        row_in = 1
        pbar = tqdm(total=ws_in.max_row)
        while row_in < ws_in.max_row:
            if not ws_in.cell(row_in, 1).value or 'Класс' not in ws_in.cell(row_in, 1).value:
                pbar.update(1)
                row_in += 1
                continue
        this_class = ws_in.cell(row_in, 1).value.split(' - ')[1]


    def search_teacher_window_by_lesson_n(self, wb, teacher_name, day_n_0, lesson_n):
        # ws
        pass


    def create_common_teacher_schedule(self, school_wb):
        ELEMENTARY_SCHOOL_TEACHERS = {'Балахонова Е. М.', 'Горбачева Е. В.', 'Домашенкина О. В.', 'Киселева Н. И.',
                                      'Стражева Г. Н.', 'Чаркина О. В.', 'Ченцова Е. Н.', 'Даймичева Р. Ф.', 'Тихоненкова А. Н.',
                                      'Смагина М. А.', 'Хретинина А. А.', 'Доронкина Л. В.', 'Мазина О. А.', 'Саватеева Г. А.',
                                      'Соколова Я. А.'}
        MAX_COL_INPUT_FILE = 111
        LESSONS_N = 11
        MAX_COL_OUTPUT_FILE = LESSONS_N * len(DAYS) + 3
        wb_out = Workbook()
        ws_out = wb_out.active
        for i in range(8):
            ws_out.append([None for i in range(MAX_COL_OUTPUT_FILE)])
        ws_out.merge_cells(start_row=1, start_column=1, end_row=6, end_column=MAX_COL_OUTPUT_FILE)
        ws_out.cell(1, 1).value = 'Расписание уроков на 2024-2025'
        ws_out.merge_cells(start_row=7, start_column=2, end_row=8, end_column=2)
        ws_out.cell(7, 2).value = 'Ф.И.О.'
        for i in range(5):
            ws_out.merge_cells(start_row=7, start_column=3 + i * 11, end_row=7, end_column=3 + (i + 1) * 11 - 1)
            ws_out.cell(7, 3 + i * 11).value = DAYS[i]
            for j in range(11):
                ws_out.cell(8, 3 + i * 11 + j).value = j + 1
        ws = school_wb.active
        input_file_row = 8
        pbar = tqdm(total=ws.max_row - input_file_row + 1)
        teacher_counter = 1
        while input_file_row <= ws.max_row:
            cur_row = [teacher_counter]
            teacher = str(ws.cell(input_file_row, 1).value)
            # if teacher in ELEMENTARY_SCHOOL_TEACHERS or teacher == 'None':
            if teacher == 'None':
                input_file_row += 1
                pbar.update(1)
                continue
            cur_row.append(teacher)
            input_file_col = 2
            while input_file_col < MAX_COL_INPUT_FILE:
                if ws.cell(input_file_row, input_file_col).value:
                    this_class = str(ws.cell(input_file_row, input_file_col).value)
                    if '_' in this_class:
                        this_class = this_class.split('_')[0]
                    this_room = str(ws.cell(input_file_row, input_file_col + 1).value)
                    if 'С' in this_room:
                        this_room = 'СЗ'
                    if "П" in this_room:
                        this_room = 'П'
                    cur_row.append('\n'.join((this_class, this_room)))
                else:
                    cur_row.append(None)
                input_file_col += 2
            cur_row.append(teacher_counter)
            ws_out.append(cur_row)
            input_file_row += 1
            teacher_counter += 1
            pbar.update(1)
        pbar.close()

        ws_out.append([None for x in range(MAX_COL_OUTPUT_FILE)])
        for output_file_col in range(3, MAX_COL_OUTPUT_FILE):
            free_cabinets = CABINETS_WITH_EL_SCHOOL[:]
            for output_file_row in range(9, 9 + teacher_counter - 1):
                if this_cell := ws_out.cell(output_file_row, output_file_col).value:
                    cabinet = this_cell.split('\n')[1]
                    if '(' in cabinet:
                        cabinet = cabinet[:-3]
                    if cabinet in free_cabinets:
                        free_cabinets.remove(cabinet)
                    else:
                        print('EXCEPTION WHILE SEARCHING FREE CABINETS!')
                        print(output_file_row, output_file_col, cabinet)
                        print(free_cabinets)
            ws_out.cell(9 + teacher_counter - 1, output_file_col).value = '\n'.join(free_cabinets)

        return wb_out

    def create_common_pupils_schedule(self, normalized_wb):
        N_CLASS = 30
        MAX_COL = N_CLASS * 2 + 4

        def create_frame():
            for i in range(56):
                ws_out.append([None for x in range(MAX_COL)])
            for i in range(1, 6):
                ws_out.merge_cells(start_row=i, start_column=1, end_row=i, end_column=11)
            for i in range(5):
                ws_out.merge_cells(start_row=7 + i * 10, start_column=1, end_row=7 + (i + 1) * 10 - 2, end_column=1)
                ws_out.cell(7 + i * 10, 1).alignment = Alignment(textRotation=90)
                ws_out.cell(7 + i * 10, 1).value = DAYS[i]
                ws_out.merge_cells(start_row=7 + i * 10, start_column=MAX_COL, end_row=7 + (i + 1) * 10 - 2, end_column=MAX_COL)
                ws_out.cell(7 + i * 10, 1).alignment = Alignment(textRotation=90)
                ws_out.cell(7 + i * 10, MAX_COL).value = DAYS[i]
                for j in range(9):
                    ws_out.cell(7 + i * 10 + j, 2).value = j + 1
                    ws_out.cell(7 + i * 10 + j, MAX_COL - 1).value = j + 1
            ws_out.cell(6, 1).alignment = Alignment(textRotation=90)
            ws_out.cell(6, 1).value = 'День'
            ws_out.cell(6, 2).alignment = Alignment(textRotation=90)
            ws_out.cell(6, 2).value = 'Урок'
            ws_out.cell(6, MAX_COL).alignment = Alignment(textRotation=90)
            ws_out.cell(6, MAX_COL).value = 'День'
            ws_out.cell(6, MAX_COL - 1).alignment = Alignment(textRotation=90)
            ws_out.cell(6, MAX_COL - 1).value = 'Урок'

        def copying_middle_school():
            nonlocal class_counter, row_in
            ws_out.cell(6, 3 + class_counter * 2).value = this_class
            row_in += 3
            pbar.update(3)
            lesson_counter = 0
            while ws_in.cell(row_in, 1).value:

                for day_counter in range(5):
                    lesson = ws_in.cell(row_in, 2 + day_counter * 2).value
                    cabinet = str(ws_in.cell(row_in, 2 + day_counter * 2 + 1).value)
                    if lesson:
                        if 'С' in cabinet:
                            cabinet = 'СЗ'
                        if "П" in cabinet:
                            cabinet = 'П'
                        if "А" in cabinet:
                            cabinet = 'АЗ'
                        ws_out.cell(7 + day_counter * 10 + lesson_counter, 3 + class_counter * 2).value = lesson
                        ws_out.cell(7 + day_counter * 10 + lesson_counter, 3 + class_counter * 2 + 1).value = cabinet
                lesson_counter += 1
                pbar.update(1)
                row_in += 1
                continue
            class_counter += 1

        def merging_high_school():
            nonlocal class_counter, row_in
            # выход через return если во входящем файле след.класс  другой
            while True:
                this_class = ws_in.cell(row_in, 1).value.split(' - ')[1]
                ws_out.cell(6, 3 + class_counter * 2).value = this_class.split('_')[0]
                row_in += 3
                pbar.update(3)
                lesson_counter = 0
                while ws_in.cell(row_in, 1).value:
                    for day_counter in range(5):
                        lesson = ws_in.cell(row_in, 2 + day_counter * 2).value
                        cabinet = str(ws_in.cell(row_in, 2 + day_counter * 2 + 1).value)
                        if lesson:
                            if 'С' in cabinet:
                                cabinet = 'СЗ'
                            if "П" in cabinet:
                                cabinet = 'П'
                            if "А" in cabinet:
                                cabinet = 'АЗ'

                            # учитываем, что если у групп общий предмет - его не надо дублировать и подписывать группы
                            if not ws_out.cell(7 + day_counter * 10 + lesson_counter, 3 + class_counter * 2).value:
                                ws_out.cell(7 + day_counter * 10 + lesson_counter, 3 + class_counter * 2).value = f'{this_class.split("_")[1]}-{lesson}'
                                ws_out.cell(7 + day_counter * 10 + lesson_counter, 3 + class_counter * 2 + 1).value = cabinet
                            else:
                                if ws_out.cell(7 + day_counter * 10 + lesson_counter, 3 + class_counter * 2 + 1).value != cabinet:
                                    ws_out.cell(7 + day_counter * 10 + lesson_counter, 3 + class_counter * 2).value += f'\n{this_class.split("_")[1]}-{lesson}'
                                    ws_out.cell(7 + day_counter * 10 + lesson_counter, 3 + class_counter * 2 + 1).value += f'\n{cabinet}'
                                else:
                                    ws_out.cell(7 + day_counter * 10 + lesson_counter, 3 + class_counter * 2).value = (
                                        ws_out.cell(7 + day_counter * 10 + lesson_counter, 3 + class_counter * 2).value.split('-'))[-1]

                    lesson_counter += 1
                    pbar.update(1)
                    row_in += 1
                    continue
                # проверка на выход. если класс тот же - class counter не рогаем. это столбцы в выходном файле.
                if not str(ws_in.cell(row_in + 1, 1).value).split(' - ')[-1].split('_')[0] == this_class.split('_')[0]:
                    class_counter += 1
                    return
                else:
                    row_in += 1


        wb_out = Workbook()
        ws_out = wb_out.active
        ws_in = normalized_wb.active
        create_frame()
        wb_out.save('what.xlsx')
        class_counter = 0
        pbar = tqdm(total=ws_in.max_row)
        row_in = 1
        while row_in < ws_in.max_row:
            if not ws_in.cell(row_in, 1).value or 'Класс' not in ws_in.cell(row_in, 1).value:
                pbar.update(1)
                row_in += 1
                continue
            this_class = ws_in.cell(row_in, 1).value.split(' - ')[1]
            #  в 10/11 классах надо сливать профили
            if '1' in this_class:
                merging_high_school()
            else:
                copying_middle_school()

        for col in range(3, MAX_COL):
            if col % 2 == 0:
                ws_out.column_dimensions[get_column_letter(col)].width = 38 * 0.138
            else:
                ws_out.column_dimensions[get_column_letter(col)].width = 125 * 0.138
        return wb_out



def normalization_scenario(file):
    wb_in = load_workbook(file)
    toolbox = FuncToolBox()
    wb_out = toolbox.row_normalization(wb_in)
    wb_out.save(f'{file.split(".")[0]}_NORM.xlsx')
    return wb_out

def checking_differences_scenario(file1, file2, normalized1=False, normalized2=False, save_normalized=True, day=-1):
    wb_in1 = load_workbook(file1)
    wb_in2 = load_workbook(file2)
    toolbox = FuncToolBox()
    # is_norm | norm in file | F
    # 0       |       0      | 1
    # 0       |       1      | 0
    # 1       |       0      | 0
    # 1       |       1      | 0
    # not (a v b)
    if not (normalized1 or 'NORM' in file1):
        wb_in1 = toolbox.row_normalization(wb_in1)
    if not (normalized2 or 'NORM' in file2):
        wb_in2 = toolbox.row_normalization(wb_in2)
    if save_normalized and not 'NORM' in file1:
        wb_in1.save(f'{file1.split(".")[0]}_NORM.xlsx')
    if save_normalized and not 'NORM' in file2:
        wb_in2.save(f'{file2.split(".")[0]}_NORM.xlsx')
    wb_out = toolbox.bold_difference(wb_in1, wb_in2)
    if day == -1:
        wb_out.save(f'{file2.split(".")[0]}_DIFFERS.xlsx')
        return wb_out
    else:
        pass


def printing_teachers_schedule_scenario(file):
    toolbox = FuncToolBox()
    wb = load_workbook(file)
    res = toolbox.create_common_teacher_schedule(wb)
    res.save(f'{file.split(".")[0]}_PRINT.xlsx')


def printing_pupils_schedule_scenario(file, normalized=False, save_normalized=True):
    toolbox = FuncToolBox()
    wb = load_workbook(file)
    if not (normalized or 'NORM' in file):
        wb = toolbox.row_normalization(wb)
    if save_normalized and not 'NORM' in file:
        wb.save(f'{file.split(".")[0]}_NORM.xlsx')
    res = toolbox.create_common_pupils_schedule(wb)
    res.save(f'{file.split(".")[0]}_PRINT.xlsx')




if __name__ == '__main__':
    normalization_scenario('пятница 4 10.xlsx')